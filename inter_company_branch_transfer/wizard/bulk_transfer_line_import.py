# -*- coding: utf-8 -*-
import base64
import csv
import io
import re

from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


# Canonical field -> accepted header aliases (matched case-insensitively,
# ignoring spaces/punctuation, so "Product Code / SKU", "product_code_sku",
# "SKU" etc. all resolve to the same column).
COLUMN_ALIASES = {
    'sku': [
        'productcode', 'productcodesku', 'sku', 'defaultcode',
        'internalreference', 'barcode', 'productcodeorsku',
    ],
    'qty': [
        'quantity', 'qty', 'productuomqty', 'productqty',
    ],
    'uom': [
        'uom', 'unitofmeasure', 'unit', 'productuom', 'productuomid',
    ],
}
REQUIRED_COLUMNS = ('sku', 'qty')


def _normalize(text):
    """Lowercase and strip everything except letters/digits, so header
    matching is forgiving of spacing, punctuation and casing."""
    return re.sub(r'[^a-z0-9]', '', str(text or '').lower())


class InterCompanyTransferBulkImportWizard(models.TransientModel):
    """Self-contained bulk import wizard for Inter-Company Transfer lines.
    Upload an .xlsx / .xls / .csv file with Product Code (SKU) and
    Quantity - matched live against Inventory (product.product) by
    Internal Reference or Barcode - and the lines are created on the
    transfer in one go."""
    _name = 'inter.company.transfer.bulk.import'
    _description = 'Bulk Import Inter-Company Transfer Lines'

    transfer_id = fields.Many2one(
        'inter.company.transfer', string='Transfer', required=True)
    transfer_reference = fields.Char(
        related='transfer_id.name', string='Transfer Reference', readonly=True)

    file = fields.Binary(string='Import File', required=True)
    filename = fields.Char(string='Filename')

    duplicate_strategy = fields.Selection([
        ('separate', 'Create Separate Lines'),
        ('merge', 'Merge Quantities'),
    ], string='Duplicate Product Strategy', default='separate', required=True,
        help='Separate: every row in the file becomes its own transfer line, '
             'even if the same product appears more than once.\n'
             'Merge: quantities for the same product code within this file '
             'are summed into a single transfer line.')

    # ------------------------------------------------------------------
    # Sample template
    # ------------------------------------------------------------------
    def action_download_sample_template(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_(
                'The "openpyxl" Python package is required to generate the '
                'sample template but is not installed on this server.'))

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Transfer Lines'

        headers = ['Product Code / SKU', 'Quantity', 'UoM']
        sheet.append(headers)
        for col in range(1, len(headers) + 1):
            sheet.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

        sheet.append(['DEMO-0001', 5, 'Units'])
        sheet.append(['DEMO-0002', 2.5, ''])

        notes = sheet.cell(row=5, column=1, value=(
            'Product Code / SKU and Quantity are required. UoM is optional '
            "- leave blank to use the product's default Unit of Measure."
        ))
        notes.font = openpyxl.styles.Font(italic=True, size=9)
        for col_letter, width in zip('ABC', (22, 12, 14)):
            sheet.column_dimensions[col_letter].width = width

        buf = io.BytesIO()
        wb.save(buf)
        content = base64.b64encode(buf.getvalue())

        attachment = self.env['ir.attachment'].create({
            'name': 'transfer_line_import_template.xlsx',
            'type': 'binary',
            'datas': content,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    # ------------------------------------------------------------------
    # File reading
    # ------------------------------------------------------------------
    def _read_rows(self):
        """Return a list of rows (each a list of cell values), first row
        being the header row."""
        self.ensure_one()
        if not self.filename:
            raise UserError(_('Please select a file to import.'))
        ext = self.filename.lower().rsplit('.', 1)[-1] if '.' in self.filename else ''
        raw = base64.b64decode(self.file)

        if ext == 'csv':
            text = raw.decode('utf-8-sig', errors='replace')
            try:
                dialect = csv.Sniffer().sniff(text.splitlines()[0])
            except (csv.Error, IndexError):
                dialect = csv.excel
            rows = list(csv.reader(io.StringIO(text), dialect))

        elif ext == 'xlsx':
            if not openpyxl:
                raise UserError(_(
                    'The "openpyxl" Python package is required to read '
                    '.xlsx files but is not installed on this server.'))
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            sheet = wb.worksheets[0]
            rows = [
                ['' if cell is None else cell for cell in row]
                for row in sheet.iter_rows(values_only=True)
            ]

        elif ext == 'xls':
            if not xlrd:
                raise UserError(_(
                    'The "xlrd" Python package is required to read legacy '
                    '.xls files but is not installed on this server.'))
            book = xlrd.open_workbook(file_contents=raw)
            sheet = book.sheet_by_index(0)
            rows = [
                [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                for r in range(sheet.nrows)
            ]

        else:
            raise UserError(_(
                'Unsupported file format "%s". Please upload a .xlsx, .xls, '
                'or .csv file.') % (ext or self.filename))

        rows = [r for r in rows if any(str(c).strip() for c in r)]
        if len(rows) < 2:
            raise UserError(_('The uploaded file contains no data rows.'))
        return rows

    def _build_column_map(self, header_row):
        alias_lookup = {}
        for canonical, aliases in COLUMN_ALIASES.items():
            for alias in aliases:
                alias_lookup[alias] = canonical

        col_map = {}
        for idx, cell in enumerate(header_row):
            canonical = alias_lookup.get(_normalize(cell))
            if canonical and canonical not in col_map:
                col_map[canonical] = idx

        missing = [c for c in REQUIRED_COLUMNS if c not in col_map]
        if missing:
            raise UserError(_(
                'The file is missing required column(s): %s. Download the '
                'sample template to see the expected headers.'
            ) % ', '.join(missing))
        return col_map

    # ------------------------------------------------------------------
    # Import
    # ------------------------------------------------------------------
    def action_import(self):
        self.ensure_one()
        transfer = self.transfer_id
        if not transfer.exists():
            raise UserError(_('The transfer this wizard was opened from no longer exists.'))
        if transfer.state != 'draft':
            raise UserError(_(
                'Lines can only be imported while the transfer is in Draft state.'))

        rows = self._read_rows()
        col_map = self._build_column_map(rows[0])
        data_rows = rows[1:]

        # ---- Pass 1: parse raw cell values row by row ----
        parsed = []
        for i, row in enumerate(data_rows, start=2):  # row 1 is the header
            def cell(key):
                idx = col_map.get(key)
                if idx is None or idx >= len(row):
                    return ''
                return row[idx]

            sku_val = cell('sku')
            sku = str(sku_val).strip()
            if isinstance(sku_val, float) and sku_val.is_integer():
                # Excel stores "1001" as 1001.0 - avoid "1001.0" mismatches.
                sku = str(int(sku_val))
            if not sku:
                continue  # silently skip rows with no product code at all
            parsed.append({
                'row': i,
                'sku': sku,
                'qty_raw': cell('qty'),
                'uom_raw': str(cell('uom')).strip() if 'uom' in col_map else '',
            })

        if not parsed:
            raise UserError(_('The uploaded file contains no data rows.'))

        # ---- Pass 2: batch-fetch products & UoMs for performance ----
        skus = list({p['sku'] for p in parsed})
        products = self.env['product.product'].search([
            '|', ('default_code', 'in', skus), ('barcode', 'in', skus),
        ])
        product_by_code = {}
        for p in products:
            if p.default_code:
                product_by_code.setdefault(p.default_code, p)
            if p.barcode:
                product_by_code.setdefault(p.barcode, p)

        uom_names = list({p['uom_raw'] for p in parsed if p['uom_raw']})
        uoms = self.env['uom.uom'].search([('name', 'in', uom_names)]) if uom_names else \
            self.env['uom.uom']
        uom_by_name = {u.name.strip().lower(): u for u in uoms}

        # ---- Pass 3: validate every row, collecting ALL errors ----
        errors = []
        resolved = []
        for p in parsed:
            product = product_by_code.get(p['sku'])
            if not product:
                # Case-insensitive fallback lookup for the occasional near-miss.
                product = self.env['product.product'].search([
                    '|',
                    ('default_code', '=ilike', p['sku']),
                    ('barcode', '=ilike', p['sku']),
                ], limit=1)
            if not product:
                errors.append(_(
                    "Row %(row)s: Product Code '%(sku)s' was not found in the system."
                ) % {'row': p['row'], 'sku': p['sku']})
                continue

            qty = None
            try:
                qty = float(str(p['qty_raw']).replace(',', '').strip())
            except (TypeError, ValueError):
                pass
            if qty is None or qty <= 0:
                errors.append(_(
                    "Row %(row)s: Invalid quantity '%(val)s'. Quantity must be a positive number."
                ) % {'row': p['row'], 'val': p['qty_raw']})
                continue

            uom = False
            if p['uom_raw']:
                uom = uom_by_name.get(p['uom_raw'].lower())
                # Unknown UoM name is not fatal - fall back to the product's
                # own unit rather than blocking the whole import.

            resolved.append({
                'row': p['row'],
                'product': product,
                'qty': qty,
                'uom': uom,
            })

        if errors:
            raise UserError('\n'.join(errors))

        # ---- Pass 4: apply duplicate strategy ----
        if self.duplicate_strategy == 'merge':
            merged = {}
            order_seq = []
            for r in resolved:
                key = r['product'].id
                if key not in merged:
                    merged[key] = dict(r)
                    order_seq.append(key)
                else:
                    merged[key]['qty'] += r['qty']
            resolved = [merged[k] for k in order_seq]

        # ---- Pass 5: create the lines atomically ----
        try:
            with self.env.cr.savepoint():
                vals_list = []
                for r in resolved:
                    product = r['product']
                    uom = r['uom'] or product.uom_id
                    vals_list.append({
                        'transfer_id': transfer.id,
                        'product_id': product.id,
                        'product_uom_qty': r['qty'],
                        'product_uom_id': uom.id,
                    })
                # A single batched create() lets the ORM group computations
                # and SQL inserts - much faster than one-by-one for large files.
                self.env['inter.company.transfer.line'].create(vals_list)
        except UserError:
            raise
        except Exception as exc:  # noqa: BLE001
            raise UserError(_(
                'The import could not be completed and no lines were created. '
                'Technical error: %s') % str(exc))

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'inter.company.transfer',
            'res_id': transfer.id,
            'view_mode': 'form',
            'target': 'current',
        }
